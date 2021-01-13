import openpyxl
import random
import math

book = openpyxl.load_workbook("./myfile.xlsx")
seating_data = [[]]
group = 4
seat = 2
mix = 2 #0:男女分座 1:男女混座 2:男女同桌
rand = True #男女同桌（mix=2）的情况下，是否要按规律分列

def main():
  read()
  write()
  save()

def start():
  pass

def read():
  sheet = book['Sheet1']
  boys = get_col_data(sheet, 1)
  girls = get_col_data(sheet, 2)
  students = {'boys':boys, 'girls':girls}
  arrange(students)

#安排座位
def arrange(data):
  global seating_data
  if type(data) == dict:
    boys = data.get('boys')
    girls = data.get('girls')
    if(mix != 1):
      random.shuffle(boys)
      random.shuffle(girls)
    if(mix == 0):
      seating_data = permutation([boys, girls])
    elif(mix == 1):
      boys_girls = boys + girls
      random.shuffle(boys_girls)
      seating_data = permutation([boys_girls])
    else:
      seating_data = permutation([mixing(boys, girls)])

def mixing(boys:list, girls:list):
  mix_data = []
  while(min(len(boys),len(girls))>0):
    random_gender_order(mix_data, boys, girls)
  mix_data.extend(girls) if(len(boys)<=len(girls)) else mix_data.extend(boys)
  '''
  if(len(boys)<=len(girls)):
    mix_data.extend(girls)
  else:
    mix_data.extend(boys)
  '''
  '''
  if(len(boys)<=len(girls)):
    while(len(boys)>0):
      random_gender_order(mix_data, boys, girls)
    mix_data.extend(girls)
  else:
    while (len(girls)>0):
      random_gender_order(mix_data, boys, girls)
    mix_data.extend(boys)
  '''
  return mix_data

def random_gender_order(mix_data, boys:list, girls:list):
  if(rand):
    if(random.randrange(2)==0):
      mix_data.append(boys.pop(0))
      mix_data.append(girls.pop(0))
    else:
      mix_data.append(girls.pop(0))
      mix_data.append(boys.pop(0))
  else:
    mix_data.append(boys.pop(0))
    mix_data.append(girls.pop(0))

def permutation(data:list[list]):
  stus = []
  desk = []
  for i in range(0, len(data)):
    data[i] = list(filter(None, data[i]))
    data[i] = list(filter(not_empty, data[i]))
    for j in range(0, len(data[i]), seat):
      for k in range(0, seat):
        if (j + k < len(data[i])):
          desk.append(data[i][j + k])
        else:
          break
      stus.append(list.copy(desk))
      desk.clear()

  stus.sort(key=lambda s:len(s), reverse=True)
  return stus

def not_empty(s):
  return s and s.strip()

def get_col_data(sheet,column):
  rows = sheet.max_row
  coldata = []
  for i in range(2, rows+1):
    cellval = sheet.cell(row=i, column=column).value
    coldata.append(cellval)
  return coldata

def write():
  sheet = book.create_sheet('seating')
  rows = math.ceil(len(seating_data)/group)
  column = group*seat

  for i in range(1, rows+1):
    for j in range(1, column+1, seat):
      if(len(seating_data)<=0):
        break
      desk = seating_data.pop(0)
      for k in range(0, len(desk)):
        if(len(desk)<=0):
          break
        set_cell_val(sheet, i, j+k, desk.pop(0))

def save():
  book.save('myfile.xlsx')

def set_cell_val(sheet, row, column, data):
  try:
    sheet.cell(row=row, column=column).value = data
  except:
    sheet.cell(row=row, column=column).value = 'write failed'

if __name__ == "__main__":
  main()